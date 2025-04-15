#!/usr/bin/env python3
"""
Memory Log Analysis Tools

This module provides tools for analyzing and visualizing memory allocation data
from legate-sparse. It includes functions for exporting data to CSV/Excel and
creating visualizations of memory usage patterns.

Optional Dependencies:
- pandas, matplotlib, seaborn: Required for visualizations
- openpyxl: Required for Excel export

Usage:
    from memlog_analysis import export_to_csv, visualize_allocations
    from memlog_parser import parse_memlog

    # Parse the log file
    allocations = parse_memlog('memlog.txt')
    
    # Export to CSV
    export_to_csv(allocations, 'memory_analysis.csv')
    
    # Create visualizations (requires pandas, matplotlib, seaborn)
    visualize_allocations(allocations)
"""  # noqa: W293

import csv
from typing import List

from memlog_parser import (
    TYPE_SIZES,
    BufferAllocation,
    are_similar_sizes,
    group_by_description,
    group_by_size,
)

# Optional imports with error handling
try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

try:
    import matplotlib.pyplot as plt

    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    import seaborn as sns

    SEABORN_AVAILABLE = True
except ImportError:
    SEABORN_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def export_to_csv(
    allocations: List[BufferAllocation],
    output_file: str,
    group_by: str = "description",
    unique_mb_only: bool = False,
    threshold_percent: float = 5.0,
):
    """
    Export memory allocation data to CSV file.

    Args:
        allocations: List of BufferAllocation objects
        output_file: Path to output CSV file
        group_by: Grouping method ('description' or 'size')
        unique_mb_only: If True, only export unique memory sizes
        threshold_percent: Percentage threshold for considering sizes similar
    """
    if group_by == "description":
        grouped_data = group_by_description(allocations)
    else:
        grouped_data = group_by_size(allocations)

    with open(output_file, "w", newline="") as f:
        writer = csv.writer(f)
        # Write header
        writer.writerow(
            [
                "Group",
                "Size (elements)",
                "Type",
                "Size (bytes)",
                "Memory (MB)",
                "File",
                "Line",
                "Timestamp",
            ]
        )

        for group, allocs in grouped_data.items():
            # Track seen entries and memory sizes for this group
            seen_entries = set()
            seen_mb_sizes = set()

            for alloc in allocs:
                mb_size = alloc.total_mb()

                # If unique_mb_only is enabled, check for similar memory sizes
                if unique_mb_only:
                    is_similar = any(
                        are_similar_sizes(mb_size, seen_size, threshold_percent)
                        for seen_size in seen_mb_sizes
                    )
                    if is_similar:
                        continue
                    seen_mb_sizes.add(mb_size)

                # Create a unique key for this entry
                entry_key = (alloc.size, alloc.type, alloc.file, alloc.line)

                # Skip if we've seen this exact entry before
                if entry_key in seen_entries:
                    continue
                seen_entries.add(entry_key)

                writer.writerow(
                    [
                        group,
                        alloc.size,
                        alloc.type,
                        TYPE_SIZES.get(alloc.type, 1),
                        mb_size,
                        alloc.file,
                        alloc.line,
                        alloc.timestamp,
                    ]
                )


def export_to_excel(allocations: List[BufferAllocation], output_file: str) -> bool:
    """
    Export memory allocation data to formatted Excel file.

    Args:
        allocations: List of BufferAllocation objects
        output_file: Path to output Excel file

    Returns:
        bool: True if export was successful, False if openpyxl is not available
    """
    if not OPENPYXL_AVAILABLE:
        print(
            "Error: Excel export requires openpyxl. Please install it with: pip install openpyxl"
        )
        return False

    wb = Workbook()

    # Create description-based view
    desc_sheet = wb.active
    desc_sheet.title = "By Description"
    _add_allocations_to_sheet(desc_sheet, allocations, group_by="description")

    # Create size-based view
    size_sheet = wb.create_sheet("By Size")
    _add_allocations_to_sheet(size_sheet, allocations, group_by="size")

    # Create summary sheet
    summary_sheet = wb.create_sheet("Summary")
    _add_summary_to_sheet(summary_sheet, allocations)

    wb.save(output_file)
    return True


def _add_allocations_to_sheet(
    sheet, allocations: List[BufferAllocation], group_by: str
):
    """Helper function to add allocations to Excel sheet with formatting."""
    if not OPENPYXL_AVAILABLE:
        return

    # Add headers
    headers = [
        "Group",
        "Size (elements)",
        "Type",
        "Size (bytes)",
        "Memory (MB)",
        "File",
        "Line",
        "Timestamp",
    ]
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    # Group data
    if group_by == "description":
        grouped_data = group_by_description(allocations)
    else:
        grouped_data = group_by_size(allocations)

    # Add data
    row = 2
    for group, allocs in grouped_data.items():
        for alloc in allocs:
            sheet.cell(row=row, column=1).value = group
            sheet.cell(row=row, column=2).value = alloc.size
            sheet.cell(row=row, column=3).value = alloc.type
            sheet.cell(row=row, column=4).value = TYPE_SIZES.get(alloc.type, 1)
            sheet.cell(row=row, column=5).value = alloc.total_mb()
            sheet.cell(row=row, column=6).value = alloc.file
            sheet.cell(row=row, column=7).value = alloc.line
            sheet.cell(row=row, column=8).value = alloc.timestamp
            row += 1

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].auto_size = True


def _add_summary_to_sheet(sheet, allocations: List[BufferAllocation]):
    """Helper function to add summary statistics to Excel sheet."""
    if not OPENPYXL_AVAILABLE:
        return

    # Calculate summary statistics
    total_memory = sum(alloc.total_mb() for alloc in allocations)
    by_type = {}
    by_description = {}

    for alloc in allocations:
        by_type[alloc.type] = by_type.get(alloc.type, 0) + alloc.total_mb()
        by_description[alloc.description] = (
            by_description.get(alloc.description, 0) + alloc.total_mb()
        )

    # Add headers
    sheet.cell(row=1, column=1).value = "Summary Statistics"
    sheet.cell(row=1, column=1).font = Font(bold=True, size=14)

    # Add total memory
    sheet.cell(row=3, column=1).value = "Total Memory Allocated:"
    sheet.cell(row=3, column=2).value = f"{total_memory:.2f} MB"

    # Add memory by type
    sheet.cell(row=5, column=1).value = "Memory by Type"
    sheet.cell(row=5, column=1).font = Font(bold=True)
    row = 6
    for type_name, memory in by_type.items():
        sheet.cell(row=row, column=1).value = type_name
        sheet.cell(row=row, column=2).value = f"{memory:.2f} MB"
        row += 1

    # Add memory by description
    sheet.cell(row=row + 1, column=1).value = "Memory by Description"
    sheet.cell(row=row + 1, column=1).font = Font(bold=True)
    row += 2
    for desc, memory in by_description.items():
        sheet.cell(row=row, column=1).value = desc
        sheet.cell(row=row, column=2).value = f"{memory:.2f} MB"
        row += 1

    # Auto-adjust column widths
    sheet.column_dimensions["A"].auto_size = True
    sheet.column_dimensions["B"].auto_size = True


def visualize_allocations(
    allocations: List[BufferAllocation],
    output_dir: str = ".",
    unique_mb_only: bool = False,
    threshold_percent: float = 5.0,
) -> bool:
    """
    Create visualizations of memory allocations.

    Args:
        allocations: List of BufferAllocation objects
        output_dir: Directory to save visualization files
        unique_mb_only: If True, only show unique memory sizes
        threshold_percent: Percentage threshold for considering sizes similar

    Returns:
        bool: True if visualization was successful, False if required libraries are not available
    """
    if not all([PANDAS_AVAILABLE, MATPLOTLIB_AVAILABLE, SEABORN_AVAILABLE]):
        print("Error: Visualization requires pandas, matplotlib, and seaborn.")
        print("Please install them with: pip install pandas matplotlib seaborn")
        return False

    # Convert to DataFrame
    data = []
    for alloc in allocations:
        data.append(
            {
                "Description": alloc.description,
                "Size_MB": alloc.total_mb(),
                "Type": alloc.type,
                "File": alloc.file,
                "Timestamp": alloc.timestamp,
            }
        )
    df = pd.DataFrame(data)

    # Create figure
    fig = plt.figure(figsize=(15, 10))  # noqa: F841

    if unique_mb_only:
        # Get unique descriptions
        unique_descriptions = set()
        for alloc in allocations:
            unique_descriptions.add(alloc.description)

        # Get unique memory sizes for each description using our similarity check
        unique_sizes = {}
        for desc in sorted(unique_descriptions):
            desc_df = df[df["Description"] == desc]
            sizes = []
            seen_sizes = set()

            for _, row in desc_df.iterrows():
                size = row["Size_MB"]
                if not any(
                    are_similar_sizes(size, seen_size, threshold_percent)
                    for seen_size in seen_sizes
                ):
                    sizes.append(size)
                    seen_sizes.add(size)

            unique_sizes[desc] = sizes

        # Create a figure with subplots for each description
        n_descriptions = len(unique_sizes)
        n_cols = min(3, n_descriptions)  # Maximum 3 columns
        n_rows = (n_descriptions + n_cols - 1) // n_cols  # Ceiling division

        for i, (desc, sizes) in enumerate(unique_sizes.items()):
            plt.subplot(n_rows, n_cols, i + 1)
            plt.bar(range(len(sizes)), sizes)
            plt.title(f"{desc}\nUnique Memory Allocations")
            plt.xlabel("Allocation Index")
            plt.ylabel("Memory (MB)")
            plt.xticks(range(len(sizes)), range(1, len(sizes) + 1))

        plt.tight_layout()
    else:
        # Memory usage by description (top 10)
        plt.subplot(2, 2, 1)
        top_descriptions = df.groupby("Description")["Size_MB"].sum().nlargest(10)
        sns.barplot(x=top_descriptions.values, y=top_descriptions.index)
        plt.title("Top 10 Memory Usage by Description")
        plt.xlabel("Memory (MB)")

        # Memory distribution by type
        plt.subplot(2, 2, 2)
        type_dist = df.groupby("Type")["Size_MB"].sum()
        plt.pie(type_dist, labels=type_dist.index, autopct="%1.1f%%")
        plt.title("Memory Distribution by Type")

    plt.tight_layout()
    plt.savefig(f"{output_dir}/memory_analysis.png", dpi=300, bbox_inches="tight")
    plt.close()
    return True


def main():
    """Example usage of the analysis tools."""
    import argparse

    from memlog_parser import parse_memlog

    parser = argparse.ArgumentParser(description="Analyze memory allocation logs")
    parser.add_argument("file", help="Path to the memory log file")
    parser.add_argument(
        "--output-dir", default=".", help="Directory to save output files"
    )
    parser.add_argument(
        "--format",
        choices=["csv", "excel", "visualization", "all"],
        default="all",
        help="Output format",
    )

    args = parser.parse_args()

    # Parse the log file
    allocations = parse_memlog(args.file)

    # Create output directory if it doesn't exist
    import os

    os.makedirs(args.output_dir, exist_ok=True)

    # Generate requested outputs
    if args.format in ["csv", "all"]:
        export_to_csv(allocations, f"{args.output_dir}/memory_analysis.csv")

    if args.format in ["excel", "all"]:
        export_to_excel(allocations, f"{args.output_dir}/memory_analysis.xlsx")

    if args.format in ["visualization", "all"]:
        visualize_allocations(allocations, args.output_dir)


if __name__ == "__main__":
    main()
